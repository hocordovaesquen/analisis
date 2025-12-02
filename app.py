import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

st.set_page_config(
    page_title="BLUSH - Sistema de Retención de Clientes",
    page_icon="💇‍♀️",
    layout="wide"
)

st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #E91E63;
        text-align: center;
        padding: 1rem;
        background: linear-gradient(90deg, #FCE4EC 0%, #F8BBD0 100%);
        border-radius: 10px;
        margin-bottom: 1rem;
    }
    .stButton>button {
        background-color: #E91E63;
        color: white;
        font-weight: bold;
        border-radius: 10px;
        padding: 0.5rem 2rem;
    }
    .metric-card {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #E91E63;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .warning-card {
        background: #FFF3CD;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #FFC107;
    }
    .success-card {
        background: #D4EDDA;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #28A745;
    }
    .vip-card {
        background: #E3F2FD;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #2196F3;
    }
</style>
""", unsafe_allow_html=True)

def agrupar_estilista(nombre):
    """Agrupa estilistas según la estructura del salón"""
    nombre = str(nombre).strip()
    
    estilistas_principales = ['Julio Luna', 'Julio', 'Julio Cesar']
    estilistas_diarios = ['Jhon', 'Yuri']
    estilistas_nuevos = ['Susy']
    administracion = ['Vero', 'Veronica']
    
    # Normalizar Julio
    if any(x in nombre for x in estilistas_principales):
        return 'Julio Luna'
    elif nombre in estilistas_diarios:
        return nombre
    elif nombre in estilistas_nuevos:
        return nombre
    elif any(x in nombre for x in administracion):
        return 'Vero'
    else:
        return 'Otros'

def es_producto(nombre_item, clase):
    """Detecta si un item es producto o servicio"""
    if pd.notna(clase):
        return str(clase).upper().strip() == 'PRODUCTO'
    
    if pd.isna(nombre_item):
        return False
    
    nombre_upper = str(nombre_item).upper()
    
    palabras_producto = [
        'MASCARILLA', 'SHAMPOO', 'SHAMPO', 'ACONDICIONADOR',
        'CREMA', 'SERUM', 'AMPOLLA', 'SPRAY', 'GEL',
        'LOTION', 'REDKEN', 'LOREAL', 'TIGI', 'KERASTASE',
        'X250ML', 'X300ML', 'X500ML', 'ML', 'GR',
        'BED HEAD', 'ALL SOFT', 'FRIZZ DISMISS'
    ]
    
    for palabra in palabras_producto:
        if palabra in nombre_upper:
            return True
    
    return False

def generar_mensaje_whatsapp(nombre, estilista, dias_sin_visita, num_visitas):
    """Genera mensaje personalizado según el perfil del cliente"""
    
    nombre_corto = nombre.split()[0] if nombre else "estimado(a) cliente"
    
    if dias_sin_visita > 90:
        mensaje = f"""¡Hola {nombre_corto}! 💇‍♀️ Somos BLUSH Hair & Make-Up y te extrañamos mucho! 

Han pasado {dias_sin_visita} días desde tu última visita con {estilista} y queremos verte de nuevo ✨

🎁 OFERTA ESPECIAL PARA TI:
- 20% de descuento en tu próximo servicio
- Válido hasta fin de mes

📍 Los Olivos, Lima
📱 Escríbenos para agendar tu cita

¡{estilista} te está esperando! 💕"""
    
    elif dias_sin_visita > 60:
        mensaje = f"""Hola {nombre_corto}! 😊

{estilista} te manda saludos desde BLUSH! ✨

Hace {dias_sin_visita} días que no te vemos y ya es hora de consentirte de nuevo 💅

¿Agendamos tu cita esta semana?
🎁 Tenemos promociones especiales para ti

¡Te esperamos! 💕"""
    
    elif dias_sin_visita > 30:
        mensaje = f"""¡{nombre_corto}! 💖

{estilista} te recuerda que ya pasaron {dias_sin_visita} días desde tu última visita a BLUSH 

Es momento de volver a lucir espectacular! ✨

¿Cuándo te viene bien para tu próxima cita?

Nos vemos pronto! 😊"""
    
    else:
        mensaje = f"""¡Hola {nombre_corto}! 

Gracias por confiar en BLUSH y en {estilista} 💕

Queremos saber si quedaste satisfecha con tu último servicio y recordarte que estamos aquí para consentirte siempre que lo necesites ✨

¡Hasta pronto! 💇‍♀️"""
    
    return mensaje

def analizar_retencion(df):
    """Analiza patrones de retención de clientes"""
    
    df['FECHA'] = pd.to_datetime(df['FECHA'], errors='coerce').ffill()
    df['MES'] = df['FECHA'].dt.to_period('M')
    df['EMPLEADO'] = df['EMPLEADO'].apply(agrupar_estilista)
    
    # Detectar productos vs servicios
    df['ES_PRODUCTO'] = df.apply(
        lambda row: es_producto(row['PRODUCTO / SERVICIO'], row.get('CLASE', None)),
        axis=1
    )
    
    hoy = datetime.now()
    
    # Análisis por cliente
    clientes = df.groupby('CLIENTE').agg({
        'FECHA': ['min', 'max', 'count'],
        'TOTAL': 'sum',
        'EMPLEADO': lambda x: x.mode()[0] if len(x.mode()) > 0 else x.iloc[0],
        'TELEF': 'first'
    }).reset_index()
    
    clientes.columns = ['CLIENTE', 'PRIMERA_VISITA', 'ULTIMA_VISITA', 'NUM_VISITAS', 'GASTO_TOTAL', 'ESTILISTA', 'TELEFONO']
    
    clientes['DIAS_SIN_VISITA'] = (hoy - pd.to_datetime(clientes['ULTIMA_VISITA'])).dt.days
    clientes['GASTO_PROMEDIO'] = clientes['GASTO_TOTAL'] / clientes['NUM_VISITAS']
    
    # Segmentación
    def segmentar_cliente(row):
        if row['NUM_VISITAS'] == 1:
            if row['DIAS_SIN_VISITA'] > 60:
                return 'Perdido'
            else:
                return 'Nuevo'
        elif row['NUM_VISITAS'] <= 3:
            if row['DIAS_SIN_VISITA'] > 90:
                return 'En Riesgo'
            else:
                return 'Ocasional'
        elif row['NUM_VISITAS'] <= 9:
            if row['DIAS_SIN_VISITA'] > 60:
                return 'En Riesgo'
            else:
                return 'Regular'
        else:
            return 'VIP'
    
    clientes['SEGMENTO'] = clientes.apply(segmentar_cliente, axis=1)
    
    # Generar mensajes
    clientes['MENSAJE_WHATSAPP'] = clientes.apply(
        lambda row: generar_mensaje_whatsapp(
            row['CLIENTE'], 
            row['ESTILISTA'], 
            row['DIAS_SIN_VISITA'],
            row['NUM_VISITAS']
        ),
        axis=1
    )
    
    return clientes, df

def calcular_metricas_estilista(df, clientes):
    """Calcula métricas detalladas por estilista"""
    
    metricas = []
    
    # Orden de visualización
    orden_estilistas = ['Julio Luna', 'Jhon', 'Yuri', 'Susy', 'Vero', 'Otros']
    
    for emp in orden_estilistas:
        df_emp = df[df['EMPLEADO'] == emp]
        clientes_emp = clientes[clientes['ESTILISTA'] == emp]
        
        if len(df_emp) == 0:
            continue
        
        # Métricas de clientes
        total_clientes = len(clientes_emp)
        retencion = (clientes_emp['NUM_VISITAS'] > 1).sum() / total_clientes * 100 if total_clientes > 0 else 0
        activos = (clientes_emp['DIAS_SIN_VISITA'] <= 60).sum()
        en_riesgo = (clientes_emp['SEGMENTO'] == 'En Riesgo').sum()
        
        # Métricas de servicios
        servicios = df_emp[~df_emp['ES_PRODUCTO']]
        productos = df_emp[df_emp['ES_PRODUCTO']]
        
        metricas.append({
            'ESTILISTA': emp,
            'TOTAL_CLIENTES': total_clientes,
            'CLIENTES_ACTIVOS': activos,
            'TASA_RETENCION': retencion,
            'CLIENTES_EN_RIESGO': en_riesgo,
            'VISITAS_PROMEDIO': clientes_emp['NUM_VISITAS'].mean() if total_clientes > 0 else 0,
            'GASTO_PROMEDIO': clientes_emp['GASTO_PROMEDIO'].mean() if total_clientes > 0 else 0,
            'TOTAL_SERVICIOS': len(servicios),
            'TOTAL_PRODUCTOS': len(productos),
            'INGRESO_SERVICIOS': servicios['TOTAL'].sum(),
            'INGRESO_PRODUCTOS': productos['TOTAL'].sum(),
            'TICKET_PROMEDIO': df_emp['TOTAL'].mean()
        })
    
    return pd.DataFrame(metricas)

def crear_excel_whatsapp(clientes_filtrados):
    """Crea Excel con lista de WhatsApp"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista WhatsApp"
    
    # Estilos
    header_fill = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = f'LISTA WHATSAPP - BLUSH SALON - {datetime.now().strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['CLIENTE', 'TELEFONO', 'ESTILISTA', 'DIAS SIN VISITA', 'SEGMENTO', 'MENSAJE']
    
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col)
        c.value = h
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = border
    
    # Datos
    fila = 4
    for _, row in clientes_filtrados.iterrows():
        ws.cell(fila, 1, row['CLIENTE'])
        ws.cell(fila, 2, str(row['TELEFONO']))
        ws.cell(fila, 3, row['ESTILISTA'])
        ws.cell(fila, 4, row['DIAS_SIN_VISITA'])
        ws.cell(fila, 5, row['SEGMENTO'])
        ws.cell(fila, 6, row['MENSAJE_WHATSAPP'])
        
        for col in range(1, 7):
            c = ws.cell(fila, col)
            c.border = border
            if col == 6:
                c.alignment = Alignment(wrap_text=True, vertical='top')
        
        fila += 1
    
    # Anchos
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 80
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# HEADER
st.markdown('<div class="main-header">💇‍♀️ BLUSH - Sistema de Retención de Clientes</div>', unsafe_allow_html=True)

# SIDEBAR
with st.sidebar:
    st.markdown("""
    <div style='text-align: center; padding: 20px; background: linear-gradient(135deg, #E91E63 0%, #9C27B0 100%); border-radius: 10px;'>
        <h1 style='color: white; margin: 0;'>💇‍♀️</h1>
        <h2 style='color: white; margin: 10px 0;'>BLUSH</h2>
        <p style='color: white; margin: 0;'>Sistema de Retención</p>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("---")
    st.markdown("### 👥 Nuestro Equipo")
    st.markdown("""
    **⭐ Estilista Principal:**
    - Julio Luna (2-4 veces/mes)
    
    **💼 Estilistas Diarios:**
    - Jhon
    - Yuri
    
    **🌱 En Desarrollo:**
    - Susy
    
    **📋 Administración:**
    - Vero
    """)
    
    st.markdown("---")
    st.markdown("### 📊 Segmentos de Clientes")
    st.markdown("""
    **🌟 VIP** - 10 o más visitas  
    Cliente muy fiel. Prioridad máxima.
    
    **💚 Regular** - 4 a 9 visitas, visitó <60 días  
    Cliente frecuente. Mantenerlo satisfecho.
    
    **💛 Ocasional** - 2 a 3 visitas  
    Viene de vez en cuando. Incentivar más visitas.
    
    **⚠️ En Riesgo** - 2+ visitas, pero >60 días sin venir  
    Puede abandonar. ¡Contactar urgente!
    
    **🆕 Nuevo** - 1 visita hace menos de 60 días  
    Primera vez reciente. Darle seguimiento.
    
    **❌ Perdido** - 1 visita hace más de 60 días  
    No regresó. Intentar reactivación.
    """)
    
    st.markdown("---")
    st.markdown("### 💡 Conceptos Clave")
    st.markdown("""
    **Tasa de Retención:**  
    % de clientes que regresaron al menos 1 vez más.  
    Se calcula: (Clientes con 2+ visitas ÷ Total) × 100  
    *Industria promedio: 15-30%*
    
    **Clientes Activos:**  
    Visitaron en los últimos 60 días (2 meses).  
    Son tu base de ingresos actual.
    
    **Clientes en Riesgo:**  
    Tienen historial pero dejaron de venir.  
    ¡La prioridad es recuperarlos!
    
    **Días sin visita:**  
    Tiempo desde su última cita.  
    Ideal: menos de 45 días.
    """)

# UPLOAD
uploaded_file = st.file_uploader(
    "📤 Sube tu archivo histórico de ventas",
    type=['xlsx', 'xls'],
    help="El archivo debe tener el formato del sistema de registro de ventas"
)

if uploaded_file:
    try:
        with st.spinner('⏳ Analizando datos...'):
            df = pd.read_excel(uploaded_file, sheet_name='Hoja1', skiprows=9)
            df = df[df['EMPLEADO'].notna()].copy()
            clientes, df_procesado = analizar_retencion(df)
            metricas_estilistas = calcular_metricas_estilista(df_procesado, clientes)
        
        st.success('✅ Análisis completado!')
        
        # KPIs PRINCIPALES
        col1, col2, col3, col4 = st.columns(4)
        
        total_clientes = len(clientes)
        tasa_retencion_global = (clientes['NUM_VISITAS'] > 1).sum() / total_clientes * 100
        clientes_riesgo = (clientes['SEGMENTO'] == 'En Riesgo').sum()
        clientes_activos = (clientes['DIAS_SIN_VISITA'] <= 60).sum()
        
        with col1:
            st.metric("👥 Total Clientes", f"{total_clientes}", 
                     help="Número total de clientes únicos que han visitado el salón")
        with col2:
            st.metric("📊 Tasa Retención", f"{tasa_retencion_global:.1f}%",
                     help="Porcentaje de clientes que regresaron (tienen 2 o más visitas). Fórmula: (Clientes con 2+ visitas ÷ Total clientes) × 100")
        with col3:
            st.metric("⚠️ En Riesgo", f"{clientes_riesgo}",
                     help="Clientes con historial (2+ visitas) que NO han visitado en los últimos 60 días. ¡Contáctalos urgente!")
        with col4:
            st.metric("✅ Activos", f"{clientes_activos}",
                     help="Clientes que visitaron en los últimos 60 días. Son tu base actual de ingresos")
        
        st.markdown("---")
        
        # TABS
        tab1, tab2, tab3, tab4 = st.tabs([
            "📊 Análisis por Estilista", 
            "👥 Segmentación", 
            "📱 Mensajes WhatsApp",
            "📈 Estadísticas Generales"
        ])
        
        with tab1:
            st.markdown("### 📊 Desempeño Completo por Estilista")
            
            st.info("💡 **Cómo leer la tabla:** La Tasa de Retención muestra qué % de clientes de cada estilista regresó (2+ visitas). Promedio de la industria: 15-30%. Verde = Mejor desempeño.")
            
            # Tabla completa con métricas
            display_metricas = metricas_estilistas[[
                'ESTILISTA', 'TOTAL_CLIENTES', 'CLIENTES_ACTIVOS', 'TASA_RETENCION',
                'TOTAL_SERVICIOS', 'TOTAL_PRODUCTOS', 'INGRESO_SERVICIOS', 'INGRESO_PRODUCTOS'
            ]].copy()
            
            display_metricas['INGRESO_TOTAL'] = display_metricas['INGRESO_SERVICIOS'] + display_metricas['INGRESO_PRODUCTOS']
            
            st.dataframe(
                display_metricas.style.format({
                    'TASA_RETENCION': '{:.1f}%',
                    'INGRESO_SERVICIOS': 'S/ {:.0f}',
                    'INGRESO_PRODUCTOS': 'S/ {:.0f}',
                    'INGRESO_TOTAL': 'S/ {:.0f}'
                }).background_gradient(cmap='RdYlGn', subset=['TASA_RETENCION']),
                use_container_width=True,
                height=300
            )
            
            st.markdown("---")
            
            # Análisis detallado por estilista principal
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("#### 👥 Detalle de Clientes por Estilista")
                
                for _, row in metricas_estilistas.iterrows():
                    emoji = "⭐" if row['ESTILISTA'] == 'Julio Luna' else "💼" if row['ESTILISTA'] in ['Jhon', 'Yuri'] else "🌱" if row['ESTILISTA'] == 'Susy' else "📋" if row['ESTILISTA'] == 'Vero' else "👤"
                    
                    color = "vip-card" if row['TASA_RETENCION'] >= 25 else "success-card" if row['TASA_RETENCION'] >= 15 else "warning-card"
                    
                    st.markdown(f"""
                    <div class='{color}'>
                        <h4>{emoji} {row['ESTILISTA']}</h4>
                        <strong>Clientes:</strong> {row['TOTAL_CLIENTES']} total | {row['CLIENTES_ACTIVOS']} activos<br>
                        <strong>Retención:</strong> {row['TASA_RETENCION']:.1f}% | Visitas/cliente: {row['VISITAS_PROMEDIO']:.1f}<br>
                        <strong>En riesgo:</strong> {row['CLIENTES_EN_RIESGO']} clientes
                    </div>
                    """, unsafe_allow_html=True)
                    st.markdown("")
            
            with col2:
                st.markdown("#### 💰 Servicios y Productos")
                
                for _, row in metricas_estilistas.iterrows():
                    emoji = "⭐" if row['ESTILISTA'] == 'Julio Luna' else "💼" if row['ESTILISTA'] in ['Jhon', 'Yuri'] else "🌱" if row['ESTILISTA'] == 'Susy' else "📋" if row['ESTILISTA'] == 'Vero' else "👤"
                    
                    st.markdown(f"""
                    <div class='metric-card'>
                        <h4>{emoji} {row['ESTILISTA']}</h4>
                        <strong>Servicios:</strong> {row['TOTAL_SERVICIOS']} (S/ {row['INGRESO_SERVICIOS']:.0f})<br>
                        <strong>Productos:</strong> {row['TOTAL_PRODUCTOS']} (S/ {row['INGRESO_PRODUCTOS']:.0f})<br>
                        <strong>Ticket promedio:</strong> S/ {row['TICKET_PROMEDIO']:.2f}
                    </div>
                    """, unsafe_allow_html=True)
                    st.markdown("")
            
            st.markdown("---")
            
            # Top clientes por estilista
            st.markdown("### 🏆 Top 5 Clientes por Estilista")
            
            for estilista in ['Julio Luna', 'Jhon', 'Yuri', 'Susy', 'Vero']:
                if estilista in clientes['ESTILISTA'].values:
                    with st.expander(f"👤 {estilista} - Top 5 Clientes"):
                        top_clientes = clientes[clientes['ESTILISTA'] == estilista].nlargest(5, 'NUM_VISITAS')[[
                            'CLIENTE', 'NUM_VISITAS', 'GASTO_TOTAL', 'DIAS_SIN_VISITA', 'SEGMENTO'
                        ]]
                        st.dataframe(
                            top_clientes.style.format({'GASTO_TOTAL': 'S/ {:.2f}'}),
                            use_container_width=True
                        )
        
        with tab2:
            st.markdown("### 👥 Segmentación de Clientes")
            
            st.info("💡 Los segmentos clasifican a tus clientes según su comportamiento de visitas. Cada color representa una acción diferente que debes tomar.")
            
            # Distribución por segmento
            segmentos = clientes['SEGMENTO'].value_counts()
            
            col1, col2 = st.columns([1, 2])
            
            with col1:
                st.markdown("#### Distribución General")
                for seg, count in segmentos.items():
                    pct = count / len(clientes) * 100
                    st.metric(seg, f"{count} ({pct:.1f}%)")
            
            with col2:
                st.markdown("#### Por Estilista Principal")
                
                # Filtrar solo estilistas principales
                estilistas_principales = ['Julio Luna', 'Jhon', 'Yuri', 'Susy', 'Vero']
                clientes_principales = clientes[clientes['ESTILISTA'].isin(estilistas_principales)]
                
                seg_estilista = pd.crosstab(clientes_principales['ESTILISTA'], clientes_principales['SEGMENTO'])
                st.dataframe(seg_estilista, use_container_width=True, height=250)
        
        with tab3:
            st.markdown("### 📱 Mensajes Personalizados para WhatsApp")
            
            st.info("💡 Filtra los clientes que quieres contactar y descarga la lista con mensajes personalizados")
            
            # Filtros
            col1, col2, col3 = st.columns(3)
            
            with col1:
                segmento_filtro = st.multiselect(
                    'Segmento',
                    options=clientes['SEGMENTO'].unique(),
                    default=['En Riesgo', 'Perdido']
                )
            
            with col2:
                estilistas_disponibles = ['Julio Luna', 'Jhon', 'Yuri', 'Susy', 'Vero', 'Otros']
                estilistas_disponibles = [e for e in estilistas_disponibles if e in clientes['ESTILISTA'].unique()]
                
                estilista_filtro = st.multiselect(
                    'Estilista',
                    options=estilistas_disponibles,
                    default=estilistas_disponibles
                )
            
            with col3:
                dias_min = st.number_input('Días mínimos sin visita', min_value=0, value=30)
            
            # Filtrar
            clientes_filtrados = clientes[
                (clientes['SEGMENTO'].isin(segmento_filtro)) &
                (clientes['ESTILISTA'].isin(estilista_filtro)) &
                (clientes['DIAS_SIN_VISITA'] >= dias_min)
            ].sort_values('DIAS_SIN_VISITA', ascending=False)
            
            st.markdown(f"#### 📋 Clientes a contactar: **{len(clientes_filtrados)}**")
            
            if len(clientes_filtrados) > 0:
                # Mostrar preview
                for idx, row in clientes_filtrados.head(5).iterrows():
                    with st.expander(f"📱 {row['CLIENTE']} - {row['ESTILISTA']}"):
                        col1, col2 = st.columns([1, 3])
                        
                        with col1:
                            st.markdown(f"""
                            **Teléfono:** {row['TELEFONO']}  
                            **Días sin visita:** {row['DIAS_SIN_VISITA']}  
                            **Visitas totales:** {row['NUM_VISITAS']}  
                            **Segmento:** {row['SEGMENTO']}
                            """)
                        
                        with col2:
                            st.markdown("**Mensaje sugerido:**")
                            st.text_area("", value=row['MENSAJE_WHATSAPP'], height=200, key=f"msg_{idx}")
                            st.markdown(f"[📱 Abrir WhatsApp](https://wa.me/51{row['TELEFONO']})")
                
                if len(clientes_filtrados) > 5:
                    st.info(f"Mostrando 5 de {len(clientes_filtrados)} clientes. Descarga el Excel para ver todos.")
                
                # Botón descarga
                st.markdown("---")
                
                excel_data = crear_excel_whatsapp(clientes_filtrados)
                
                col1, col2, col3 = st.columns([1, 2, 1])
                with col2:
                    st.download_button(
                        label=f"📥 DESCARGAR LISTA COMPLETA ({len(clientes_filtrados)} clientes)",
                        data=excel_data,
                        file_name=f"WhatsApp_BLUSH_{datetime.now().strftime('%d%m%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            else:
                st.warning("No hay clientes que cumplan con los filtros seleccionados")
        
        with tab4:
            st.markdown("### 📈 Estadísticas Generales")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("#### 📊 Resumen de Visitas")
                st.markdown(f"""
                - **Promedio de visitas por cliente:** {clientes['NUM_VISITAS'].mean():.2f}
                - **Mediana de visitas:** {clientes['NUM_VISITAS'].median():.0f}
                - **Cliente más frecuente:** {clientes['NUM_VISITAS'].max():.0f} visitas
                - **Clientes con 1 sola visita:** {(clientes['NUM_VISITAS'] == 1).sum()} ({(clientes['NUM_VISITAS'] == 1).sum()/len(clientes)*100:.1f}%)
                """)
            
            with col2:
                st.markdown("#### 💰 Análisis de Gasto")
                st.markdown(f"""
                - **Gasto promedio por visita:** S/ {clientes['GASTO_PROMEDIO'].mean():.2f}
                - **Gasto total promedio por cliente:** S/ {clientes['GASTO_TOTAL'].mean():.2f}
                - **Cliente con mayor gasto:** S/ {clientes['GASTO_TOTAL'].max():.2f}
                """)
            
            st.markdown("---")
            st.markdown("#### 🎯 Top 10 Clientes VIP del Salón")
            
            top_vip = clientes.nlargest(10, 'NUM_VISITAS')[[
                'CLIENTE', 'NUM_VISITAS', 'GASTO_TOTAL', 'ESTILISTA', 'DIAS_SIN_VISITA'
            ]]
            
            st.dataframe(
                top_vip.style.format({
                    'GASTO_TOTAL': 'S/ {:.2f}'
                }),
                use_container_width=True
            )
    
    except Exception as e:
        st.error(f"❌ Error al procesar el archivo: {str(e)}")
        st.info("Verifica que el archivo tenga el formato correcto")

else:
    st.info("👆 Sube tu archivo histórico de ventas para comenzar el análisis")
    
    st.markdown("---")
    
    # Explicación educativa de conceptos
    st.markdown("### 📚 Guía del Sistema - Conceptos Explicados")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        #### 📊 ¿Qué es la Tasa de Retención?
        
        Es el **porcentaje de clientes que regresan** al salón después de su primera visita.
        
        **Fórmula:**
        ```
        Retención = (Clientes con 2+ visitas ÷ Total clientes) × 100
        ```
        
        **Ejemplo:**
        - Tienes 100 clientes totales
        - 25 han regresado (tienen 2+ visitas)
        - Tu retención es: 25%
        
        **¿Es bueno o malo?**
        - 🔴 <15% = Problema serio
        - 🟡 15-20% = Por debajo del promedio
        - 🟢 20-30% = Promedio de la industria  
        - 🔵 30%+ = ¡Excelente!
        
        **¿Por qué importa?**  
        Retener un cliente cuesta 5 veces menos que conseguir uno nuevo. Un cliente que regresa gastará 3-5 veces más en su vida.
        """)
        
        st.markdown("---")
        
        st.markdown("""
        #### 💰 Otros Conceptos Importantes
        
        **Clientes Activos:**  
        Los que visitaron en los últimos 60 días. Son tu flujo de caja actual.
        
        **Clientes en Riesgo:**  
        Tienen 2+ visitas pero no vienen hace 60+ días. ¡Están a punto de perderse!
        
        **Días sin visita:**  
        Cuánto tiempo pasó desde su última cita. Lo ideal es <30 días para servicios de belleza.
        
        **Ticket Promedio:**  
        Cuánto gasta cada cliente por visita en promedio.
        """)
    
    with col2:
        st.markdown("""
        #### 🎯 Segmentos de Clientes Explicados
        
        | Segmento | Visitas | Última visita | ¿Qué hacer? |
        |----------|---------|---------------|-------------|
        | 🌟 **VIP** | 10+ | Cualquiera | Recompensar. Son tus mejores clientes. |
        | 💚 **Regular** | 4-9 | <60 días | Mantener satisfechos. Base estable. |
        | 💛 **Ocasional** | 2-3 | <90 días | Incentivar más visitas frecuentes. |
        | ⚠️ **En Riesgo** | 2+ | >60 días | **¡Contactar urgente!** Pueden irse. |
        | 🆕 **Nuevo** | 1 | <60 días | Dar seguimiento. Potencial de retención. |
        | ❌ **Perdido** | 1 | >60 días | Reactivar con promoción especial. |
        
        ---
        
        #### 📱 ¿Para qué sirve el WhatsApp?
        
        El sistema genera **mensajes personalizados** para cada cliente según:
        - Cuántos días sin visitar
        - Cuántas veces ha venido
        - Su estilista preferido
        - Su segmento
        
        **Ejemplo de mensaje generado:**
        ```
        ¡Hola María! 😊
        
        Yuri te manda saludos desde BLUSH! ✨
        
        Hace 65 días que no te vemos y ya 
        es hora de consentirte de nuevo 💅
        
        ¿Agendamos tu cita esta semana?
        
        ¡Te esperamos! 💕
        ```
        
        Solo copias, pegas y envías por WhatsApp.
        """)
    
    st.markdown("---")
    st.markdown("### 🎯 ¿Qué hace esta herramienta?")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        #### 📊 Analiza
        - Desempeño por estilista
        - Clientes y servicios
        - Patrones de retención
        - Productos vs servicios
        """)
    
    with col2:
        st.markdown("""
        #### 🎯 Identifica
        - Clientes en riesgo
        - Oportunidades de reactivación
        - Top clientes por estilista
        - Áreas de mejora
        """)
    
    with col3:
        st.markdown("""
        #### 📱 Genera
        - Mensajes WhatsApp personalizados
        - Listas de contactos por estilista
        - Reportes descargables
        - Acciones concretas
        """)

# FOOTER
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666;'>
    <p>💇‍♀️ <b>BLUSH Hair & Make-Up Salon</b> | Los Olivos, Lima</p>
    <p style='font-size: 0.8rem;'>Sistema de Retención de Clientes v2.5 - Conceptos Explicados</p>
</div>
""", unsafe_allow_html=True)
